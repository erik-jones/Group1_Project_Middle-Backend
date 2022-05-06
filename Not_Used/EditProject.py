
import openpyxl

# Constants
ELDB_FILE_PATH = "../../Data/Example_DB.xlsx"

#to open the workbook
wb_obj = openpyxl.load_workbook(ELDB_FILE_PATH)

# get workbook active sheet object from the active attribute
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column

# field is the column that needs to change, value is the new value you want to set it to
def editProject(projectName, field, value):
    
    # find the project within the database (find the row that's needed)
    for i in range(2, max_row + 1):
        row_obj = sheet_obj.cell(row = i, column = 1)
        if row_obj.value == projectName:
            # identify which column needs to change
            for j in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = 1, column = j)
                if cell_obj.value == field:
                    # using the row (i) and column (j) found in previous steps, change the field to the new value
                    sheet_obj.cell(row = i, column = j).value = value
                    
    #overwrite existing database file (save)
    wb_obj.save("../Data/Example_DB.xlsx")
