# Loops through the columns of the excel database, builds a dictionary where key is column name and value is its respective column number
# Ex. columnTranslate = {"columnName1": 1, "columnName2": 2}
import openpyxl

def translateColumns(filepath):
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook["Sheet1"]

    colNums = {}
    current = 1
    for col in worksheet.iter_cols(1, worksheet.max_column):
        colNums[col[0].value] = current
        current += 1

    return(colNums)
