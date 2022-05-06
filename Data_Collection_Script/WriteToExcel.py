# Writes the data from each table to their respective excel sheet
import openpyxl
from openpyxl import Workbook
import ColumnTranslate

# Constants
GROUP_SHEET = "Subgroup"
IMPACT_SHEET = "Impact"
EVENT_SHEET = "Event"
COURSE_SHEET = "Course"

# Return the max number of rows that have data so we can append to the end
def __getMaxRows(sheet):
    rows = 0
    for max_row, row in enumerate(sheet, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

# Write data to respective sheet
def __writeToSheet(db: Workbook, data, columnNums, sheetName):
    columns = columnNums
    sheet = db[sheetName]

    # Add a check to see if this is already in the excel file
    
    # Get the next empty row in the sheet
    newRow = __getMaxRows(sheet) + 1

    for field in data:
        if field in columns:
            # If item is a list, you have to add it differently
            if isinstance(data.get(field), list):
                fieldAsString = ""
                counter = 0
                for val in data.get(field):
                    fieldAsString = fieldAsString + ", " + data.get(field)[counter]
                    counter += 1
                sheet.cell(row=newRow, column=columns.get(field)).value = data.get(fieldAsString)
            else:    
                sheet.cell(row=newRow, column=columns.get(field)).value = data.get(field)

# Main function of this script
def writeData(groupData, impactData, eventData, courseData, excelFilePath):
    db = openpyxl.load_workbook(excelFilePath)
    
    print("Writing groups to excel")
    for group in groupData:
        __writeToSheet(db, group, ColumnTranslate.groupColumns, GROUP_SHEET)

    print("Writing impacts to excel")
    for impact in impactData:
        __writeToSheet(db, impact, ColumnTranslate.impactColumns, IMPACT_SHEET)

    print("Writing events to excel")
    for event in eventData:
        __writeToSheet(db, event, ColumnTranslate.eventColumns, EVENT_SHEET)

    # print("Writing courses to exel")
    # for course in courseData:
    #     __writeToSheet(db, event, ColumnTranslate.courseColumns, COURSE_SHEET)

    db.save(excelFilePath)
